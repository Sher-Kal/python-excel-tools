# excel_fill_placeholders.py
import re
import json
import sys
from openpyxl import load_workbook

PLACEHOLDER = re.compile(r"\{\{(\w+)\}\}")

def fill_placeholders(template_xlsx='template.xlsx', data_json='data.json', out_xlsx='filled.xlsx'):
    with open(data_json, 'r', encoding='utf-8') as f:
        data = json.load(f)  # e.g., {"name":"Ivan", "date":"2025-09-01"}

    wb = load_workbook(template_xlsx)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if isinstance(val, str):
                    def repl(m):
                        key = m.group(1)
                        return str(data.get(key, m.group(0)))
                    new_val = PLACEHOLDER.sub(repl, val)
                    if new_val != val:
                        cell.value = new_val
    wb.save(out_xlsx)
    print(f"OK: wrote {out_xlsx}")

if __name__ == '__main__':
    args = sys.argv[1:]
    fill_placeholders(*args)  # template_xlsx, data_json, out_xlsx
